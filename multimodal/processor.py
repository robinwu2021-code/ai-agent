"""
multimodal/processor.py — 多模态输入处理器

将各种文件格式统一转为 LLM 可消费的消息内容。
支持：图像 / PDF / Word / Excel / CSV / 音频（stub）
"""
from __future__ import annotations

import base64
import csv
import io
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import structlog

log = structlog.get_logger(__name__)


@dataclass
class ProcessedInput:
    """处理后的输入内容，可直接附加到 Message.content。"""
    type:     str   # "text" | "image" | "table" | "audio_transcript"
    text:     str   # 文本表示（所有类型都有）
    raw:      Any = None   # 原始数据（图像的 base64 等）
    mime:     str = ""
    filename: str = ""


class MultimodalProcessor:
    """
    多模态文件处理器。
    将上传文件转换为 ProcessedInput 列表，注入到 AgentInput。
    """

    IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".gif", ".webp"}
    AUDIO_EXTS = {".mp3", ".wav", ".m4a", ".ogg"}

    def __init__(self, llm_engine: Any | None = None) -> None:
        self._llm = llm_engine  # 用于图像描述、音频转文字

    async def process_file(
        self, path: str | Path, hint: str = ""
    ) -> ProcessedInput:
        """根据文件扩展名自动选择处理策略。"""
        p      = Path(path)
        suffix = p.suffix.lower()
        name   = p.name

        if suffix in self.IMAGE_EXTS:
            return await self._process_image(p, hint)
        elif suffix == ".pdf":
            return self._process_pdf(p)
        elif suffix in (".docx", ".doc"):
            return self._process_docx(p)
        elif suffix in (".xlsx", ".xls"):
            return self._process_excel(p)
        elif suffix == ".csv":
            return self._process_csv(p)
        elif suffix in (".json",):
            return self._process_json(p)
        elif suffix in self.AUDIO_EXTS:
            return await self._process_audio(p)
        else:
            # 默认读文本
            text = p.read_text(errors="replace")[:8000]
            return ProcessedInput(type="text", text=text, filename=name)

    async def process_url_image(self, url: str) -> ProcessedInput:
        """从 URL 下载并处理图像。"""
        try:
            import httpx
            async with httpx.AsyncClient() as client:
                resp  = await client.get(url, timeout=10)
                data  = base64.b64encode(resp.content).decode()
                ctype = resp.headers.get("content-type", "image/jpeg")
            return ProcessedInput(
                type="image", text=f"[图像来自 URL: {url}]",
                raw=data, mime=ctype,
            )
        except Exception as e:
            return ProcessedInput(type="text", text=f"[图像加载失败: {e}]")

    # ── Private processors ─────────────────────────────────────────

    async def _process_image(self, path: Path, hint: str) -> ProcessedInput:
        data     = base64.b64encode(path.read_bytes()).decode()
        suffix   = path.suffix.lower().lstrip(".")
        mime_map = {"jpg": "image/jpeg", "jpeg": "image/jpeg",
                    "png": "image/png", "gif": "image/gif", "webp": "image/webp"}
        mime     = mime_map.get(suffix, "image/jpeg")

        # 如有 LLM Engine 则自动生成描述（供无视觉 LLM 使用）
        description = f"[图像文件: {path.name}]"
        if self._llm and hasattr(self._llm, "describe_image"):
            try:
                description = await self._llm.describe_image(data, mime, hint)
            except Exception:
                pass

        log.debug("multimodal.image", file=path.name, bytes=len(path.read_bytes()))
        return ProcessedInput(
            type="image", text=description,
            raw=data, mime=mime, filename=path.name,
        )

    def _process_pdf(self, path: Path) -> ProcessedInput:
        try:
            import pypdf
            reader = pypdf.PdfReader(str(path))
            pages  = [p.extract_text() or "" for p in reader.pages]
            text   = "\n\n".join(pages)[:12000]
            log.debug("multimodal.pdf", file=path.name, pages=len(pages))
            return ProcessedInput(
                type="text",
                text=f"[PDF: {path.name}, {len(reader.pages)} 页]\n\n{text}",
                filename=path.name,
            )
        except ImportError:
            return ProcessedInput(
                type="text",
                text=f"[PDF 解析需要 pypdf: pip install pypdf]",
                filename=path.name,
            )

    def _process_docx(self, path: Path) -> ProcessedInput:
        try:
            import docx
            doc  = docx.Document(str(path))
            text = "\n".join(p.text for p in doc.paragraphs)[:10000]
            return ProcessedInput(
                type="text",
                text=f"[Word 文档: {path.name}]\n\n{text}",
                filename=path.name,
            )
        except ImportError:
            return ProcessedInput(
                type="text",
                text=f"[DOCX 解析需要 python-docx: pip install python-docx]",
                filename=path.name,
            )

    def _process_excel(self, path: Path) -> ProcessedInput:
        try:
            import openpyxl
            wb     = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
            sheets = []
            for name in wb.sheetnames[:3]:   # 最多处理 3 个 Sheet
                ws   = wb[name]
                rows = []
                for row in ws.iter_rows(max_row=50, values_only=True):
                    rows.append("\t".join(str(c or "") for c in row))
                sheets.append(f"### Sheet: {name}\n" + "\n".join(rows))
            text = "\n\n".join(sheets)
            return ProcessedInput(
                type="table",
                text=f"[Excel: {path.name}]\n\n{text}",
                filename=path.name,
            )
        except ImportError:
            return ProcessedInput(
                type="text",
                text=f"[Excel 解析需要 openpyxl: pip install openpyxl]",
                filename=path.name,
            )

    def _process_csv(self, path: Path) -> ProcessedInput:
        with open(path, newline="", encoding="utf-8", errors="replace") as f:
            reader  = csv.reader(f)
            rows    = [row for i, row in enumerate(reader) if i < 100]
        preview = "\n".join("\t".join(r) for r in rows)
        return ProcessedInput(
            type="table",
            text=f"[CSV: {path.name}, 前 {len(rows)} 行]\n\n{preview}",
            filename=path.name,
        )

    def _process_json(self, path: Path) -> ProcessedInput:
        raw  = path.read_text(errors="replace")
        try:
            data = json.loads(raw)
            text = json.dumps(data, ensure_ascii=False, indent=2)[:6000]
        except Exception:
            text = raw[:6000]
        return ProcessedInput(
            type="text",
            text=f"[JSON: {path.name}]\n\n{text}",
            filename=path.name,
        )

    async def _process_audio(self, path: Path) -> ProcessedInput:
        # Stub：生产接入 Whisper API
        log.warning("multimodal.audio.stub", file=path.name)
        return ProcessedInput(
            type="audio_transcript",
            text=f"[音频文件: {path.name}（生产环境请接入 Whisper API 转写）]",
            filename=path.name,
        )

    # ── Batch processing ──────────────────────────────────────────

    async def process_files(
        self, files: list[dict]
    ) -> list[ProcessedInput]:
        """
        批量处理 AgentInput.files 列表。
        files 格式：[{"name": "x.pdf", "path": "/tmp/x.pdf"}] 或 {"url": "..."}
        """
        results = []
        for f in files:
            path = f.get("path") or f.get("url", "")
            if not path:
                continue
            try:
                if path.startswith("http"):
                    result = await self.process_url_image(path)
                else:
                    result = await self.process_file(path)
                results.append(result)
            except Exception as e:
                log.error("multimodal.error", file=path, error=str(e))
                results.append(ProcessedInput(
                    type="text", text=f"[处理失败: {path}: {e}]"
                ))
        return results

    @staticmethod
    def to_context_text(inputs: list[ProcessedInput]) -> str:
        """将处理结果转为可注入 Prompt 的文本。"""
        if not inputs:
            return ""
        parts = ["\n## 附件内容\n"]
        for pi in inputs:
            parts.append(f"**{pi.filename or pi.type}**\n{pi.text}\n")
        return "\n".join(parts)
